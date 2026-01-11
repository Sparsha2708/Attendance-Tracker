import openpyxl
wb= openpyxl.load_workbook(r"ATTENDANCE.xlsx")
sheet= wb.active
wb1=openpyxl.Workbook()#create workbook object
Sheet=wb1.active#get active worksheet
Sheet.title="low attendance"#set title of worksheet
Sheet["A1"].value="S NO"    
Sheet["B1"].value="NAME"
Sheet["C1"].value="USN"
Sheet["D1"].value="ATTENDANCE"

j=2

for i in range(2, sheet.max_row+1):
    if(int(sheet["D"+str(i)].value)<75):
        Sheet["D"+str(j)].value=sheet["D"+str(i)].value
        Sheet["A"+str(j)].value=j-1
        Sheet["B"+str(j)].value=sheet["B"+str(i)].value
        Sheet["C"+str(j)].value=sheet["C"+str(i)].value
        j=j+1
        wb1.save("low_attendance.xlsx")
wb.close()
wb1.close()
'''l1=[]
l2=[]
rows = list(Sheet.iter_rows(values_only=True))

for i in range(len(rows)):
    row1=(rows[i])
    l1.append(len(str(row1)))
    for j in range(0,3):
        print(str(row1[j]).ljust(max(l1)),end="")
    print()'''

'''col_widths = []
for col in zip(*rows):
    col_widths.append(max(len(str(cell)) for cell in col))

for row in rows:
    for i, cell in enumerate(row):
        print(str(cell).ljust(col_widths[i] + 3), end="")
    print()
'''

